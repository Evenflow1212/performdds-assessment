"""
Perform DDS — Production Extractor
Reads ADA code PDF (Procedures by Provider) and populates
the standard 9-tab Perform DDS Assessment Workbook.

Usage:
    python production_extractor.py input.pdf template.xlsx output.xlsx --months 12
"""

import re
import sys
import argparse
import pdfplumber
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict
from copy import copy

# ─────────────────────────────────────────────────────────────────────────────
# CODE MAPPING: ADA code → Perform DDS group
# ─────────────────────────────────────────────────────────────────────────────

CODE_MAP = {
    # ── EXAMS ─────────────────────────────────────────────────────────────────
    "D0120": "exam_periodic",
    "D0140": "exam_focused",
    "D0145": "exam_other",
    "D0150": "exam_comprehensive",
    "D0170": "exam_other",
    "D0180": "exam_perio",

    # ── IMAGING ───────────────────────────────────────────────────────────────
    "D0210": "imaging_fmx",
    "D0220": "imaging_other",
    "D0230": "imaging_other",
    "D0270": "imaging_other",
    "D0272": "imaging_other",
    "D0273": "imaging_other",
    "D0274": "imaging_bw4",
    "D0330": "imaging_pano",
    "D0340": "imaging_other",
    "D0350": "imaging_other",
    "D0364": "imaging_other",
    "D0365": "imaging_other",
    "D0367": "imaging_other",
    "D0380": "imaging_other",

    # ── HYGIENE ───────────────────────────────────────────────────────────────
    "D1110": "hyg_adult_prophy",
    "D1120": "hyg_child_prophy",
    "D1206": "hyg_other",
    "D1351": "hyg_other",
    "D1353": "hyg_other",
    "D1354": "hyg_other",
    "D1556": "hyg_other",
    "D4341": "hyg_srp",
    "D4342": "hyg_srp",
    "D4346": "hyg_irrigation",
    "D4381": "hyg_arestin",
    "D4910": "hyg_perio_maint",
    "D4921": "hyg_irrigation",

    # ── CROWN & BRIDGE (left-side rows) ───────────────────────────────────────
    "D2740": "cb_2740",
    "D2750": "cb_2750",

    # ── OTHER CROWNS (right-side orange tables) ────────────────────────────────
    "D2962": "oc_veneers",
    "D2780": "oc_cast_metal",
    "D2790": "oc_cast_metal",
    "D2799": "oc_interim",
    "D2719": "oc_crown_prep",
    "D6057": "oc_custom_abutment",
    "D6058": "oc_implant_crown",
    "D6059": "oc_implant_crown",
    "D6060": "oc_implant_crown",
    "D6061": "oc_implant_crown",
    "D6062": "oc_implant_crown",
    "D6063": "oc_implant_crown",
    "D6064": "oc_implant_crown",
    "D6065": "oc_implant_crown",

    # BRIDGE UNITS
    "D6245": "bridge_pontic",
    "D6740": "bridge_pontic",
    "D6253": "bridge_other",
    "D6930": "bridge_other",
    "D6999": "bridge_other",

    # INLAYS/ONLAYS (grouped into Crown & Bridge "other" on left side)
    "D2510": "cb_inlay_onlay",
    "D2520": "cb_inlay_onlay",
    "D2530": "cb_inlay_onlay",
    "D2542": "cb_inlay_onlay",
    "D2543": "cb_inlay_onlay",
    "D2544": "cb_inlay_onlay",
    "D2610": "cb_inlay_onlay",
    "D2620": "cb_inlay_onlay",
    "D2630": "cb_inlay_onlay",
    "D2642": "cb_inlay_onlay",
    "D2643": "cb_inlay_onlay",
    "D2644": "cb_inlay_onlay",
    "D2650": "cb_inlay_onlay",
    "D2651": "cb_inlay_onlay",
    "D2652": "cb_inlay_onlay",
    "D2662": "cb_inlay_onlay",
    "D2663": "cb_inlay_onlay",
    "D2664": "cb_inlay_onlay",

    # ORTHO
    "D8090": "ortho_comp_adult",
    "D8040": "ortho_limited",
    "D8681": "ortho_retention",
    "D8682": "ortho_retention",
    "D8695": "ortho_retention",
    "D8696": "ortho_retention",
    "D8699": "ortho_retention",
    "D8701": "ortho_retention",
    "D8702": "ortho_retention",
    "D8703": "ortho_retention",
    "D8704": "ortho_retention",

    # DENTURES
    "D5110": "den_complete",
    "D5120": "den_complete",
    "D5130": "den_complete",
    "D5140": "den_complete",
    "D5213": "den_partial_cast",
    "D5214": "den_partial_cast",
    "D5211": "den_partial_resin",
    "D5223": "den_partial_resin",
    "D5225": "den_partial_flex",
    "D5226": "den_partial_flex",
    "D5820": "den_interim",
    "D5821": "den_interim",
    "D5410": "den_adj",
    "D5411": "den_adj",
    "D5421": "den_adj",
    "D5422": "den_adj",
    "D5511": "den_repair",
    "D5512": "den_repair",
    "D5520": "den_repair",
    "D5612": "den_repair",
    "D5640": "den_repair",
    "D5650": "den_repair",
    "D5730": "den_repair",
    "D5741": "den_repair",
    "D5750": "den_repair",
    "D5751": "den_repair",
    "D5851": "den_repair",
    "D5862": "den_repair",
    "D6111": "den_implant_supported",
    "D6112": "den_implant_supported",
    "D6113": "den_implant_supported",

    # ENDO
    "D3310": "endo_anterior",
    "D3320": "endo_bicuspid",
    "D3330": "endo_molar",
    "D3332": "endo_retreat",
    "D3346": "endo_retreat",
    "D3347": "endo_retreat",
    "D3348": "endo_retreat",
    "D3110": "endo_other",
    "D3120": "endo_other",
    "D3240": "endo_other",

    # PERIO
    "D4249": "perio_crown_length",
    "D4211": "perio_gingivectomy",
    "D4212": "perio_gingivectomy",
    "D4921": "perio_irrigation",

    # ORAL SURGERY
    "D7922": "os_membrane",
    "D7953": "os_bone_graft_ridge",
    "D6104": "os_bone_graft_implant",
    "D7140": "os_ext_simple",
    "D7210": "os_ext_surgical",
    "D6010": "os_implant_surgical",
    "D6011": "os_implant_misc",
    "D6080": "os_implant_misc",
    "D6089": "os_implant_misc",
    "D6090": "os_implant_misc",
    "D6091": "os_implant_misc",
    "D6092": "os_implant_misc",
    "D6100": "os_implant_misc",
    "D6192": "os_implant_misc",
    "D6197": "os_implant_misc",
    "D6198": "os_implant_misc",
    "D6199": "os_implant_misc",
    "D7465": "os_abscess",
    "D7510": "os_abscess",
    "D7511": "os_abscess",
    "D7971": "os_abscess",
}

# Groups that roll into "inlays & onlays; veneers & other" left-side row
INLAY_ONLAY_GROUPS = {"cb_inlay_onlay", "oc_veneers"}

# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSING
# ─────────────────────────────────────────────────────────────────────────────

def normalize_code(code: str) -> str:
    """Normalize ADA codes to 5-char format: D120 → D0120, D274 → D0274."""
    code = code.strip().upper()
    if not code.startswith("D"):
        code = "D" + code
    digits = code[1:]
    if len(digits) == 3:
        digits = "0" + digits
    return "D" + digits


GENERIC_SKIP_PATTERNS = (
    r'---',
    r'TIME ',
    r'DATE ',
    r'PROCEDURES BY',
    r'From EOD',
    r'Service$',
    r'Number$',
    r'Amount$',
    r'Page \d',
    r'Current Dental',
    r'© American',
    r'\*$',
)


def parse_pdf(pdf_path: str, skip_pattern: str = None) -> dict:
    """
    Parse Eaglesoft 'Procedures by Provider' PDF.

    Eaglesoft format (per provider section):
        qty
        (blank)
        $amount
        Description
        00120        <- 5-digit ADA code, no D prefix

    `skip_pattern` is an optional regex (alternation allowed, e.g. "Foo|Bar")
    appended to the built-in GENERIC_SKIP_PATTERNS for client-specific
    header/footer lines.

    Returns dict: { "D0120": {"qty": N, "total": $, "avg": $}, ... }
    """
    import pdfplumber

    def clean_num(s):
        return float(s.replace(',', '').replace('$', '').strip()) if s else 0.0

    code_re   = re.compile(r'^(\d{5}[A-Z]?)$')
    amount_re = re.compile(r'^\$?([\d,]+\.?\d+)$')
    qty_re    = re.compile(r'^(\d+)$')

    patterns = list(GENERIC_SKIP_PATTERNS)
    if skip_pattern:
        patterns.append(skip_pattern)
    skip_re = re.compile(r'^(' + '|'.join(patterns) + ')')

    raw_lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            raw_lines.extend(text.split('\n'))

    # Normalize lines (preserve blanks for positional logic)
    lines = [l.strip() for l in raw_lines]

    codes = defaultdict(lambda: {'qty': 0, 'total': 0.0})

    for i, line in enumerate(lines):
        if not code_re.match(line):
            continue

        raw_code = line  # e.g. "00120"
        # Normalize: strip leading zero → D0120
        norm = 'D' + raw_code.lstrip('0') if raw_code.lstrip('0') else 'D0'
        # But keep 4 digits minimum: D0120 not D120
        digits = raw_code[1:]  # skip first zero
        if len(digits) < 3:
            digits = raw_code  # edge case
        norm = 'D0' + digits if len(digits) == 3 else 'D' + digits

        # Scan backwards for qty and amount (blank lines between them)
        qty, amt = 0, 0.0
        found_amt, found_qty = False, False

        for j in range(i - 1, max(0, i - 10) - 1, -1):
            l = lines[j]
            if not l:
                continue  # skip blanks
            if skip_re.match(l):
                break
            if code_re.match(l):
                break  # hit another code

            if not found_amt and amount_re.match(l):
                amt = clean_num(l)
                found_amt = True
                continue

            if found_amt and not found_qty and qty_re.match(l):
                qty = int(l)
                found_qty = True
                break  # got both

        if found_qty and found_amt and qty > 0:
            codes[norm]['qty']   += qty
            codes[norm]['total'] += amt

    # Finalize with averages
    result = {}
    for code, data in codes.items():
        q = data['qty']
        t = round(data['total'], 2)
        result[code] = {'qty': q, 'total': t, 'avg': round(t / q, 2) if q else 0.0}

    return result


# ─────────────────────────────────────────────────────────────────────────────
# AGGREGATION: flatten raw codes → Perform DDS groups
# ─────────────────────────────────────────────────────────────────────────────

def aggregate(raw: dict) -> dict:
    """
    Aggregate raw code data into Perform DDS groups.
    Returns dict: { group_key: {"qty": N, "total": $, "avg": $} }
    """
    groups = defaultdict(lambda: {"qty": 0, "total": 0.0})

    for code, data in raw.items():
        group = CODE_MAP.get(normalize_code(code))
        if group:
            groups[group]["qty"] += data["qty"]
            groups[group]["total"] += data["total"]

    # Compute averages
    result = {}
    for group, data in groups.items():
        qty = data["qty"]
        total = data["total"]
        avg = round(total / qty, 2) if qty > 0 else 0.0
        result[group] = {"qty": qty, "total": round(total, 2), "avg": avg}

    return result


def get_group(g: dict, key: str, field: str = "qty"):
    return g.get(key, {}).get(field, 0) or 0


# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK POPULATION
# ─────────────────────────────────────────────────────────────────────────────

def populate_workbook(template_path: str, output_path: str, raw: dict, groups: dict,
                      months: int, practice_name: str = ""):
    wb = load_workbook(template_path)
    pop_production_worksheet(wb, groups, raw, months, practice_name)
    pop_all_codes(wb, raw)
    wb.save(output_path)
    print(f"✅  Saved: {output_path}")


def set_val(ws, cell_ref, value):
    ws[cell_ref] = value


def pop_production_worksheet(wb, g: dict, raw: dict, months: int, practice_name: str):
    ws = wb["Production Worksheet"]

    # ── HEADER ────────────────────────────────────────────────────────────────
    if practice_name:
        ws["D4"] = practice_name
    ws["D5"] = months

    # Total production = sum all raw totals
    total_prod = sum(v["total"] for v in raw.values())
    ws["G5"] = round(total_prod, 2)

    # ── EXAMS ─────────────────────────────────────────────────────────────────
    ws["D10"] = get_group(g, "exam_periodic", "qty")
    ws["D11"] = get_group(g, "exam_focused", "qty")
    ws["D12"] = get_group(g, "exam_comprehensive", "qty")
    ws["D13"] = get_group(g, "exam_perio", "qty")

    # per-month formulas
    ws["E10"] = f"=D10/D5"
    ws["E11"] = f"=D11/D5"
    ws["E12"] = f"=D12/D5"
    ws["E13"] = f"=D13/D5"

    # avg fees from raw
    ws["F10"] = raw.get("D0120", {}).get("avg", 0)
    ws["F11"] = raw.get("D0140", {}).get("avg", 0)
    ws["F12"] = raw.get("D0150", {}).get("avg", 0)
    ws["F13"] = raw.get("D0180", {}).get("avg", 0)

    # ── IMAGING ───────────────────────────────────────────────────────────────
    ws["D16"] = get_group(g, "imaging_fmx", "qty")
    ws["D17"] = get_group(g, "imaging_bw4", "qty")
    ws["D18"] = get_group(g, "imaging_pano", "qty")

    ws["E16"] = "=D16/D5"
    ws["E17"] = "=D17/D5"
    ws["E18"] = "=D18/D5"

    ws["F16"] = raw.get("D0210", {}).get("avg", 0)
    ws["F17"] = raw.get("D0274", {}).get("avg", 0)
    ws["F18"] = raw.get("D0330", {}).get("avg", 0)

    # ── HYGIENE ───────────────────────────────────────────────────────────────
    adult_prophy_qty = get_group(g, "hyg_adult_prophy", "qty")
    child_prophy_qty = get_group(g, "hyg_child_prophy", "qty")
    perio_maint_qty  = get_group(g, "hyg_perio_maint", "qty")
    srp_qty          = get_group(g, "hyg_srp", "qty")
    arestin_qty      = get_group(g, "hyg_arestin", "qty")
    irrigation_qty   = get_group(g, "hyg_irrigation", "qty")

    ws["D21"] = adult_prophy_qty
    ws["D22"] = child_prophy_qty
    ws["D23"] = perio_maint_qty
    ws["D24"] = srp_qty
    ws["D25"] = arestin_qty
    ws["D26"] = irrigation_qty

    ws["E21"] = "=D21/D5"
    ws["E22"] = "=D22/D5"
    ws["E23"] = "=D23/D5"
    ws["E24"] = "=D24/D5"
    ws["E25"] = "=D25/D5"
    ws["E26"] = "=D26/D5"

    ws["F21"] = raw.get("D1110", {}).get("avg", 0)
    ws["F22"] = raw.get("D1120", {}).get("avg", 0)
    ws["F23"] = raw.get("D4910", {}).get("avg", 0)
    # SRP avg: weighted average of D4341 + D4342
    srp_total = get_group(g, "hyg_srp", "total")
    ws["F24"] = round(srp_total / srp_qty, 2) if srp_qty else 0
    ws["F25"] = raw.get("D4381", {}).get("avg", 0)
    ws["F26"] = raw.get("D4346", {}).get("avg", 0)

    ws["G21"] = "=D21*F21"
    ws["G22"] = "=D22*F22"
    ws["G23"] = "=D23*F23"
    ws["G24"] = "=D24*F24"
    ws["G25"] = "=D25*F25"
    ws["G26"] = "=D26*F26"

    # ── CROWN & BRIDGE (left side) ────────────────────────────────────────────
    crown_2740_qty = get_group(g, "cb_2740", "qty")
    crown_2750_qty = get_group(g, "cb_2750", "qty")
    inlay_onlay_qty   = get_group(g, "cb_inlay_onlay", "qty")
    inlay_onlay_total = get_group(g, "cb_inlay_onlay", "total")
    bridge_qty    = get_group(g, "bridge_pontic", "qty")
    bridge_total  = get_group(g, "bridge_pontic", "total")
    implant_c_qty   = get_group(g, "oc_implant_crown", "qty")
    implant_c_total = get_group(g, "oc_implant_crown", "total")

    ws["D31"] = crown_2740_qty
    ws["D32"] = crown_2750_qty
    ws["D33"] = inlay_onlay_qty
    ws["D34"] = bridge_qty
    ws["D35"] = implant_c_qty

    ws["E31"] = "=D31/D5"
    ws["E32"] = "=D32/D5"
    ws["E33"] = "=D33/D5"
    ws["E34"] = "=D34/D5"
    ws["E35"] = "=D35/D5"

    ws["F31"] = raw.get("D2740", {}).get("avg", 0)
    ws["F32"] = raw.get("D2750", {}).get("avg", 0)
    ws["F33"] = round(inlay_onlay_total / inlay_onlay_qty, 2) if inlay_onlay_qty else 0
    ws["F34"] = round(bridge_total / bridge_qty, 2) if bridge_qty else 0
    ws["F35"] = round(implant_c_total / implant_c_qty, 2) if implant_c_qty else 0

    ws["G31"] = "=D31*F31"
    ws["G32"] = "=D32*F32"
    ws["G33"] = "=D33*F33"
    ws["G34"] = "=D34*F34"
    ws["G35"] = "=D35*F35"

    # Average crowns/mo (all crown types)
    total_crowns = crown_2740_qty + crown_2750_qty + bridge_qty + implant_c_qty
    ws["E36"] = round(total_crowns / months, 2) if months else 0

    # ── SPECIALTY % ───────────────────────────────────────────────────────────
    perio_total  = get_group(g, "perio_crown_length", "total") + get_group(g, "perio_gingivectomy", "total")
    perio_total += get_group(g, "perio_irrigation", "total")
    os_total     = sum(get_group(g, k, "total") for k in g if k.startswith("os_"))
    ortho_total  = sum(get_group(g, k, "total") for k in g if k.startswith("ortho_"))
    endo_total   = sum(get_group(g, k, "total") for k in g if k.startswith("endo_"))
    den_total    = sum(get_group(g, k, "total") for k in g if k.startswith("den_"))

    ws["D39"] = round(perio_total, 2)
    ws["D40"] = round(os_total, 2)
    ws["D41"] = round(ortho_total, 2)
    ws["D42"] = round(endo_total, 2)
    ws["D43"] = round(den_total, 2)

    if total_prod > 0:
        ws["E39"] = "=D39/G5"
        ws["E40"] = "=D40/G5"
        ws["E41"] = "=D41/G5"
        ws["E42"] = "=D42/G5"
        ws["E43"] = "=D43/G5"

    ws["D44"] = "=SUM(D39:D43)"
    ws["E44"] = "=D44/G5"

    # ── RIGHT-SIDE ORANGE TABLES ───────────────────────────────────────────────
    _fill_right_tables(ws, g, raw)


def _fill_right_tables(ws, g: dict, raw: dict):
    """Populate J–N orange detail tables on Production Worksheet."""

    # OTHER CROWNS (rows 3–9)
    def oc(row, code_key, qty, total, avg):
        ws[f"L{row}"] = qty
        ws[f"M{row}"] = round(total, 2)
        ws[f"N{row}"] = round(avg, 2)

    oc(3,  "veneers",
       get_group(g, "oc_veneers", "qty"),
       get_group(g, "oc_veneers", "total"),
       get_group(g, "oc_veneers", "avg"))

    cast_qty   = get_group(g, "oc_cast_metal", "qty")
    cast_total = get_group(g, "oc_cast_metal", "total")
    oc(4, "cast_metal", cast_qty, cast_total,
       round(cast_total/cast_qty, 2) if cast_qty else 0)

    oc(5,  "interim",
       get_group(g, "oc_interim", "qty"),
       get_group(g, "oc_interim", "total"),
       get_group(g, "oc_interim", "avg"))

    oc(6,  "custom_abutment",
       get_group(g, "oc_custom_abutment", "qty"),
       get_group(g, "oc_custom_abutment", "total"),
       get_group(g, "oc_custom_abutment", "avg"))

    oc(7,  "crown_prep",
       get_group(g, "oc_crown_prep", "qty"),
       get_group(g, "oc_crown_prep", "total"),
       get_group(g, "oc_crown_prep", "avg"))

    oc(8,  "implant_crown",
       get_group(g, "oc_implant_crown", "qty"),
       get_group(g, "oc_implant_crown", "total"),
       get_group(g, "oc_implant_crown", "avg"))

    # Row 9 = SUM row for OTHER CROWNS (qty/total)
    ws["L9"] = "=SUM(L3:L8)"
    ws["M9"] = "=SUM(M3:M8)"
    ws["N9"] = "=IFERROR(M9/L9,0)"

    # BRIDGE UNITS (rows 12–18, SUM row 18)
    bridge_qty   = get_group(g, "bridge_pontic", "qty")
    bridge_total = get_group(g, "bridge_pontic", "total")
    ws["L12"] = bridge_qty
    ws["M12"] = round(bridge_total, 2)
    ws["N12"] = round(bridge_total/bridge_qty, 2) if bridge_qty else 0
    ws["L18"] = "=SUM(L12:L17)"
    ws["M18"] = "=SUM(M12:M17)"
    ws["N18"] = "=IFERROR(M18/L18,0)"

    # ORTHO (rows 21–28, SUM row 28)
    comp_qty   = get_group(g, "ortho_comp_adult", "qty")
    comp_total = get_group(g, "ortho_comp_adult", "total")
    ws["L21"] = comp_qty
    ws["M21"] = round(comp_total, 2)
    ws["N21"] = round(comp_total/comp_qty, 2) if comp_qty else 0

    ltd_qty   = get_group(g, "ortho_limited", "qty")
    ltd_total = get_group(g, "ortho_limited", "total")
    ws["L22"] = ltd_qty
    ws["M22"] = round(ltd_total, 2)
    ws["N22"] = round(ltd_total/ltd_qty, 2) if ltd_qty else 0

    ret_qty   = get_group(g, "ortho_retention", "qty")
    ret_total = get_group(g, "ortho_retention", "total")
    ws["L23"] = ret_qty
    ws["M23"] = round(ret_total, 2)
    ws["N23"] = round(ret_total/ret_qty, 2) if ret_qty else 0

    ws["L28"] = "=SUM(L21:L27)"
    ws["M28"] = "=SUM(M21:M27)"
    ws["N28"] = "=IFERROR(M28/L28,0)"

    # DENTURES (rows 31–40, SUM row 40)
    def den(row, key):
        qty   = get_group(g, key, "qty")
        total = get_group(g, key, "total")
        ws[f"L{row}"] = qty
        ws[f"M{row}"] = round(total, 2)
        ws[f"N{row}"] = round(total/qty, 2) if qty else 0

    den(31, "den_complete")
    den(32, "den_partial_cast")
    den(33, "den_partial_flex")
    den(34, "den_partial_resin")
    den(35, "den_interim")
    den(36, "den_implant_supported")
    den(37, "den_adj")
    den(38, "den_repair")

    ws["L40"] = "=SUM(L31:L39)"
    ws["M40"] = "=SUM(M31:M39)"
    ws["N40"] = "=IFERROR(M40/L40,0)"

    # ENDO (rows 43–47, SUM row 47)
    def endo(row, key):
        qty   = get_group(g, key, "qty")
        total = get_group(g, key, "total")
        ws[f"L{row}"] = qty
        ws[f"M{row}"] = round(total, 2)
        ws[f"N{row}"] = round(total/qty, 2) if qty else 0

    endo(43, "endo_anterior")
    endo(44, "endo_bicuspid")
    endo(45, "endo_molar")
    endo(46, "endo_retreat")

    ws["L47"] = "=SUM(L43:L46)"
    ws["M47"] = "=SUM(M43:M46)"
    ws["N47"] = "=IFERROR(M47/L47,0)"

    # PERIO (rows 51–55, SUM row 55)
    def perio(row, key):
        qty   = get_group(g, key, "qty")
        total = get_group(g, key, "total")
        ws[f"L{row}"] = qty
        ws[f"M{row}"] = round(total, 2)
        ws[f"N{row}"] = round(total/qty, 2) if qty else 0

    perio(51, "perio_crown_length")
    perio(52, "perio_gingivectomy")
    perio(56, "perio_irrigation")   # Gingival Irrigation

    ws["L55"] = "=SUM(L51:L54)"
    ws["M55"] = "=SUM(M51:M54)"
    ws["N55"] = "=IFERROR(M55/L55,0)"

    # ORAL SURGERY (rows 58–68, SUM row 68)
    def os(row, key):
        qty   = get_group(g, key, "qty")
        total = get_group(g, key, "total")
        ws[f"L{row}"] = qty
        ws[f"M{row}"] = round(total, 2)
        ws[f"N{row}"] = round(total/qty, 2) if qty else 0

    os(58, "os_membrane")
    os(59, "os_bone_graft_ridge")
    os(60, "os_bone_graft_implant")
    os(61, "os_ext_simple")
    os(62, "os_ext_surgical")
    os(63, "os_implant_surgical")
    os(65, "os_implant_misc")
    os(66, "os_abscess")

    ws["L68"] = "=SUM(L58:L67)"
    ws["M68"] = "=SUM(M58:M67)"
    ws["N68"] = "=IFERROR(M68/L68,0)"


def pop_all_codes(wb, raw: dict):
    """Populate 'All Codes - Production Report' tab with every code found."""
    ws = wb["All Codes - Production Report"]

    # Clear existing data rows (keep header row 1)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Write codes sorted
    total_prod = sum(v["total"] for v in raw.values())
    sorted_codes = sorted(raw.items(), key=lambda x: x[0])

    for i, (code, data) in enumerate(sorted_codes, start=2):
        ws[f"A{i}"] = code
        ws[f"C{i}"] = data["qty"]
        ws[f"D{i}"] = data["total"]
        ws[f"E{i}"] = data["avg"]
        ws[f"F{i}"] = round(data["total"] / total_prod, 10) if total_prod else 0


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Perform DDS Production Extractor")
    parser.add_argument("pdf",      help="Path to Procedures by Provider PDF")
    parser.add_argument("template", help="Path to Assessment Workbook template (.xlsx)")
    parser.add_argument("output",   help="Output .xlsx path")
    parser.add_argument("--months", type=int, default=12, help="Number of months reviewed (default 12)")
    parser.add_argument("--practice", default="", help="Practice name")
    parser.add_argument("--skip-pattern", default=None,
                        help="Additional regex pattern for client-specific header/footer "
                             "lines to skip during PDF parsing (e.g. 'Foo Family|JD Foo'). "
                             "Appended to built-in generic filters.")
    args = parser.parse_args()

    print(f"📄 Parsing PDF: {args.pdf}")
    raw = parse_pdf(args.pdf, skip_pattern=args.skip_pattern)
    print(f"   Found {len(raw)} ADA codes")

    groups = aggregate(raw)
    print(f"   Mapped to {len(groups)} Perform DDS groups")

    print(f"📊 Populating workbook...")
    populate_workbook(args.template, args.output, raw, groups,
                      months=args.months, practice_name=args.practice)


if __name__ == "__main__":
    main()
